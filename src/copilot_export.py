"""MODULE 10 — Export center for Analyst Copilot deliverables."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from src.analyst_models import CopilotResults
from src.export_utils import _sanitize_filename, export_docx, export_markdown, export_pdf


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def requirements_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# Catálogo de Requisitos\n", "| ID | Requisito | Categoria | Prioridade | Documento | Página |", "|---|---|---|---|---|---|"]
    for req in copilot.requirements:
        lines.append(
            f"| {req.id} | {req.requirement} | {req.category} | {req.priority} "
            f"| {req.source_document} | {req.page} |"
        )
    return "\n".join(lines) + "\n"


def user_stories_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# User Stories\n"]
    for story in copilot.user_stories:
        lines.append(f"## {story.id} ({story.requirement_id})\n")
        lines.append(story.story + "\n")
        lines.append("**Critérios de aceitação:**\n")
        for criterion in story.acceptance_criteria:
            lines.append(f"- {criterion}")
        lines.append("")
    return "\n".join(lines)


def risks_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# Registo de Riscos\n", "| ID | Risco | Categoria | Impacto | Probabilidade | Recomendação |", "|---|---|---|---|---|---|"]
    for risk in copilot.risks:
        lines.append(
            f"| {risk.id} | {risk.risk} | {risk.category} | {risk.impact} "
            f"| {risk.likelihood} | {risk.recommendation} |"
        )
    return "\n".join(lines) + "\n"


def ambiguities_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# Ambiguidades Detetadas\n"]
    for item in copilot.ambiguities:
        lines.append(f"## {item.id}\n")
        lines.append(f"**Afirmação:** {item.statement}\n")
        lines.append(f"**Problema:** {item.issue}\n")
        lines.append(f"**Pergunta sugerida:** {item.suggested_question}\n")
    return "\n".join(lines)


def gaps_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# Análise de Lacunas\n", "| ID | Lacuna | Área | Recomendação |", "|---|---|---|---|"]
    for gap in copilot.gaps:
        lines.append(f"| {gap.id} | {gap.gap} | {gap.area} | {gap.recommendation} |")
    return "\n".join(lines) + "\n"


def traceability_to_markdown(copilot: CopilotResults) -> str:
    lines = [
        "# Matriz de Rastreabilidade\n",
        "| ID | Requisito | Documento | Página | Trecho |",
        "|---|---|---|---|---|",
    ]
    for row in copilot.traceability:
        lines.append(
            f"| {row.requirement_id} | {row.requirement} | {row.document} "
            f"| {row.page} | {row.chunk_id} |"
        )
    return "\n".join(lines) + "\n"


def stakeholder_questions_to_markdown(copilot: CopilotResults) -> str:
    lines = ["# Perguntas para Stakeholders\n"]
    for index, question in enumerate(copilot.stakeholder_questions, start=1):
        lines.append(f"{index}. {question}")
    return "\n".join(lines) + "\n"


def build_copilot_full_report(copilot: CopilotResults, collection_label: str) -> str:
    """Assemble all Analyst Copilot deliverables into one report."""
    sections = [
        f"# Relatório Analyst Copilot\n\n**Coleção:** {collection_label}\n",
    ]
    if copilot.executive_summary:
        sections.append("---\n\n" + copilot.executive_summary.strip())
    sections.append("---\n\n" + requirements_to_markdown(copilot))
    if copilot.user_stories:
        sections.append("---\n\n" + user_stories_to_markdown(copilot))
    if copilot.risks:
        sections.append("---\n\n" + risks_to_markdown(copilot))
    if copilot.ambiguities:
        sections.append("---\n\n" + ambiguities_to_markdown(copilot))
    if copilot.gaps:
        sections.append("---\n\n" + gaps_to_markdown(copilot))
    if copilot.stakeholder_questions:
        sections.append("---\n\n" + stakeholder_questions_to_markdown(copilot))
    if copilot.traceability:
        sections.append("---\n\n" + traceability_to_markdown(copilot))
    return "\n".join(sections).strip() + "\n"


def export_copilot_excel(copilot: CopilotResults, title: str = "Analyst Copilot") -> tuple[bytes, str]:
    """Export all copilot deliverables to a multi-sheet Excel workbook."""
    return _build_workbook(copilot, title=title, sheets=None)


def export_single_excel(
    copilot: CopilotResults,
    report_type: str,
    title: str = "Analyst Copilot",
) -> tuple[bytes, str]:
    """Export a single deliverable as a one-sheet Excel workbook."""
    return _build_workbook(copilot, title=title, sheets=[report_type])


def _build_workbook(
    copilot: CopilotResults,
    title: str,
    sheets: list[str] | None,
) -> tuple[bytes, str]:
    """Build Excel workbook for all or selected report sheets."""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(name: str, headers: list[str], rows: list[list]) -> None:
        sheet = workbook.create_sheet(name[:31])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

    sheet_definitions: dict[str, tuple[list[str], list[list]]] = {
        "Relatório completo": ([], []),
        "Catálogo de Requisitos": (
            ["ID", "Requirement", "Category", "Priority", "Document", "Page", "Chunk"],
            [
                [r.id, r.requirement, r.category, r.priority, r.source_document, r.page, r.chunk_id]
                for r in copilot.requirements
            ],
        ),
        "User Stories": (
            ["ID", "Requirement ID", "Story", "Acceptance Criteria"],
            [
                [s.id, s.requirement_id, s.story, "; ".join(s.acceptance_criteria)]
                for s in copilot.user_stories
            ],
        ),
        "Registo de Riscos": (
            ["ID", "Risk", "Category", "Impact", "Likelihood", "Recommendation", "Document", "Page"],
            [
                [r.id, r.risk, r.category, r.impact, r.likelihood, r.recommendation, r.source_document, r.page]
                for r in copilot.risks
            ],
        ),
        "Ambiguidades": (
            ["ID", "Statement", "Issue", "Suggested Question", "Document", "Page"],
            [
                [a.id, a.statement, a.issue, a.suggested_question, a.source_document, a.page]
                for a in copilot.ambiguities
            ],
        ),
        "Análise de Lacunas": (
            ["ID", "Gap", "Area", "Recommendation"],
            [[g.id, g.gap, g.area, g.recommendation] for g in copilot.gaps],
        ),
        "Matriz de Rastreabilidade": (
            ["Requirement ID", "Requirement", "Document", "Page", "Chunk"],
            [
                [t.requirement_id, t.requirement, t.document, t.page, t.chunk_id]
                for t in copilot.traceability
            ],
        ),
        "Perguntas Stakeholders": (
            ["#", "Question"],
            [[index, question] for index, question in enumerate(copilot.stakeholder_questions, start=1)],
        ),
    }

    selected = sheets or [
        "Catálogo de Requisitos",
        "User Stories",
        "Registo de Riscos",
        "Ambiguidades",
        "Análise de Lacunas",
        "Matriz de Rastreabilidade",
        "Perguntas Stakeholders",
    ]

    for sheet_name in selected:
        if sheet_name == "Resumo Executivo" and copilot.executive_summary:
            summary_sheet = workbook.create_sheet("Executive Summary")
            for line in copilot.executive_summary.splitlines():
                summary_sheet.append([line])
            continue

        headers, rows = sheet_definitions.get(sheet_name, ([], []))
        if headers:
            add_sheet(sheet_name, headers, rows)

    if sheets is None and copilot.executive_summary:
        summary_sheet = workbook.create_sheet("Executive Summary")
        for line in copilot.executive_summary.splitlines():
            summary_sheet.append([line])

    buffer = io.BytesIO()
    workbook.save(buffer)
    export_label = sheets[0] if sheets and len(sheets) == 1 else title
    filename = f"{_sanitize_filename(export_label)}_{_timestamp()}.xlsx"
    return buffer.getvalue(), filename


def export_copilot_report(
    copilot: CopilotResults,
    collection_label: str,
    fmt: str,
) -> tuple[bytes, str]:
    """Export full copilot report in markdown, docx, pdf, or excel."""
    title = f"Analyst Copilot — {collection_label}"
    if fmt == "excel":
        return export_copilot_excel(copilot, title=title)

    report = build_copilot_full_report(copilot, collection_label)
    if fmt == "markdown":
        return export_markdown(report, title=title)
    if fmt == "docx":
        return export_docx(report, title=title)
    if fmt == "pdf":
        return export_pdf(report, title=title)
    raise ValueError(f"Formato não suportado: {fmt}")
